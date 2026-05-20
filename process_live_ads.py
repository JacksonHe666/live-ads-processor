#!/usr/bin/env python3
import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


SOURCE_HEADERS = [
    "加热主播",
    "创建时间",
    "名称",
    "编号",
    "加热开始时间",
    "加热结束时间",
    "实际加热时长（秒）",
    "总消耗",
    "直播间曝光人数",
    "直播间观看人数",
    "直播间新增粉丝数",
    "直播间评论次数",
    "直播间点赞次数",
    "直播间超过1分钟观看人数",
    "总成交金额",
    "总成交ROI",
    "总成交订单数",
    "总下单金额",
    "总下单订单数",
]

INPUT_ALIASES = {
    "当场成交金额": "总成交金额",
    "当场成交ROI": "总成交ROI",
    "当场成交订单数": "总成交订单数",
    "当场下单金额": "总下单金额",
    "当场下单订单数": "总下单订单数",
}

NUMERIC_HEADERS = {
    "实际加热时长（秒）",
    "总消耗",
    "直播间曝光人数",
    "直播间观看人数",
    "直播间新增粉丝数",
    "直播间评论次数",
    "直播间点赞次数",
    "直播间超过1分钟观看人数",
    "总成交金额",
    "总成交ROI",
    "总成交订单数",
    "总下单金额",
    "总下单订单数",
}

SUMMARY_HEADERS = [
    "投放目标",
    "总消耗（豆）",
    "总消耗金额",
    "服务费",
    "合计金额",
    "总下单金额",
    "下单roi",
    "总成交金额",
    "总成交roi",
    "总成交订单数",
    "单个成交成本",
    "曝光人数",
    "观看人数",
    "场观均价",
    "直播间进入率",
    "新增粉丝",
    "新增均价",
    "新增/场观",
    "评论次数",
    "评论均价",
    "评论/场观",
    "点赞次数",
    "互动总数",
    "超过1分钟观看人数",
    "一分钟观看均价",
]

TARGET_ROWS = ["", "控成交", "放量成交", "roi", "涨粉", "互动", "商点", "观众"]
SUMMARY_METRICS = SUMMARY_HEADERS[1:]

SETTLEMENT_HEADERS = [
    "日期场次",
    "成交金额",
    "成交roi",
    "成交订单数",
    "单个成交成本",
    "曝光人数",
    "场观",
    "场观均价",
    "曝光进入率",
    "增粉",
    "新增均价",
    "评论",
    "实际消耗微信豆",
    "实际消耗金额（元）",
    "点赞次数",
]


def clean_header(value):
    return str(value or "").replace("\ufeff", "").strip()


def to_number(value):
    if value is None:
        return 0
    text = str(value).strip()
    if text == "" or text.upper() == "N/A":
        return 0
    text = text.replace(",", "").replace("¥", "").replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return 0
    return int(number) if number.is_integer() else number


def safe_ratio(numerator, denominator):
    return numerator / denominator if denominator else 0


def normalize_date_text(value):
    text = str(value or "").strip()
    if text.upper() == "N/A":
        return ""
    match = re.match(r"^(\d{4})年0?(\d{1,2})月0?(\d{1,2})日(?:\s+\d{1,2}:\d{2})?", text)
    if match:
        year, month, day = match.groups()
        return f"{year}年{int(month)}月{int(day)}日"
    return text


def read_csv_rows(input_path):
    return read_csv_file(input_path)[1]


def read_csv_file(input_path):
    with Path(input_path).open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        raw_headers = [clean_header(header) for header in (reader.fieldnames or [])]
        rows = []
        for raw in reader:
            normalized = {}
            for key, value in raw.items():
                header = INPUT_ALIASES.get(clean_header(key), clean_header(key))
                normalized[header] = value
            rows.append(normalized)
    return raw_headers, rows


def read_csv_files(input_paths):
    all_rows = []
    header_sets = []
    for input_path in input_paths:
        raw_headers, rows = read_csv_file(input_path)
        header_sets.append((Path(input_path).name, raw_headers))
        all_rows.extend(rows)
    return header_sets, all_rows


def validate_inputs(header_sets, rows):
    warnings = []
    for filename, headers in header_sets:
        normalized_headers = {INPUT_ALIASES.get(header, header) for header in headers}
        missing = [header for header in SOURCE_HEADERS if header not in normalized_headers]
        if missing:
            warnings.append(f"{filename} 缺少表头：{'、'.join(missing)}。强制转化可能出现错误。")
    if len(header_sets) > 1:
        first_headers = [INPUT_ALIASES.get(header, header) for header in header_sets[0][1]]
        for filename, headers in header_sets[1:]:
            normalized = [INPUT_ALIASES.get(header, header) for header in headers]
            if normalized != first_headers:
                warnings.append(f"{filename} 的表头顺序或字段与第一张 CSV 不一致。")
    anchors = sorted({str(row.get("加热主播", "")).strip() for row in rows if str(row.get("加热主播", "")).strip()})
    if len(anchors) > 1:
        warnings.append(f"检测到多个加热主播：{'、'.join(anchors)}。请确认是否应合并处理。")
    return warnings


def default_output_stem(rows):
    if not rows:
        return "直播投放数据"
    date_text = ""
    for row in rows:
        date_text = normalize_date_text(row.get("加热开始时间", "")) or normalize_date_text(row.get("创建时间", ""))
        if date_text:
            break
    anchor = str(rows[0].get("加热主播", "") or "").strip()
    return f"{date_text}{anchor}直播投放数据" if date_text or anchor else "直播投放数据"


def row_category(name):
    text = str(name or "").lower()
    if "控成交" in text:
        return "控成交"
    if "放量成交" in text:
        return "放量成交"
    if "roi" in text:
        return "roi"
    for target in ["涨粉", "互动", "商点", "观众"]:
        if target in text:
            return target
    return ""


def aggregate_rows(rows):
    metrics = [
        "总消耗",
        "直播间曝光人数",
        "直播间观看人数",
        "直播间新增粉丝数",
        "直播间评论次数",
        "直播间点赞次数",
        "直播间超过1分钟观看人数",
        "总成交金额",
        "总成交订单数",
        "总下单金额",
    ]
    aggregates = {target: {metric: 0 for metric in metrics} for target in TARGET_ROWS}
    for row in rows:
        category = row_category(row.get("名称"))
        for target in {"", category}:
            if target not in aggregates:
                continue
            for metric in metrics:
                aggregates[target][metric] += to_number(row.get(metric))
    return aggregates


def summary_values(aggregates, target):
    agg = aggregates[target]
    spend_beans = agg["总消耗"]
    spend_yuan = spend_beans / 10
    order_amount = agg["总下单金额"]
    deal_amount = agg["总成交金额"]
    deal_orders = agg["总成交订单数"]
    exposure = agg["直播间曝光人数"]
    views = agg["直播间观看人数"]
    fans = agg["直播间新增粉丝数"]
    comments = agg["直播间评论次数"]
    likes = agg["直播间点赞次数"]
    one_min_views = agg["直播间超过1分钟观看人数"]
    return {
        "总消耗（豆）": spend_beans,
        "总消耗金额": spend_yuan,
        "服务费": spend_yuan * 0.1,
        "合计金额": spend_yuan * 1.1,
        "总下单金额": order_amount,
        "下单roi": safe_ratio(order_amount, spend_yuan),
        "总成交金额": deal_amount,
        "总成交roi": safe_ratio(deal_amount, spend_yuan),
        "总成交订单数": deal_orders,
        "单个成交成本": safe_ratio(spend_yuan, deal_orders),
        "曝光人数": exposure,
        "观看人数": views,
        "场观均价": safe_ratio(spend_yuan, views),
        "直播间进入率": safe_ratio(views, exposure),
        "新增粉丝": fans,
        "新增均价": safe_ratio(spend_yuan, fans),
        "新增/场观": safe_ratio(fans, views),
        "评论次数": comments,
        "评论均价": safe_ratio(spend_yuan, comments),
        "评论/场观": safe_ratio(comments, views),
        "点赞次数": likes,
        "互动总数": comments + likes,
        "超过1分钟观看人数": one_min_views,
        "一分钟观看均价": safe_ratio(spend_yuan, one_min_views),
    }


def style_sheet(ws, header_fill="1F4E78"):
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=header_fill)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def apply_summary_group_styles(ws, transpose=True):
    header_fill = PatternFill("solid", fgColor="7030A0")
    metric_fill = PatternFill("solid", fgColor="E7D9F2")
    group_fills = [
        (2, 5, "F7E6D5"),
        (6, 11, "E4F0EA"),
        (12, 15, "E5EDF8"),
        (16, 18, "F3E7F7"),
        (19, 21, "FEF0D7"),
        (22, 25, "E9E2D5"),
    ]
    thin = Side(style="thin", color="E6DFD8")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx in range(2, ws.max_row + 1):
        for start, end, color in group_fills:
            if start <= row_idx <= end:
                fill = PatternFill("solid", fgColor=color)
                break
        else:
            fill = PatternFill("solid", fgColor="FFFFFF")
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = metric_fill if col_idx == 1 else fill
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def delete_zero_summary_columns(ws):
    for col_idx in range(ws.max_column, 2, -1):
        values = [to_number(ws.cell(row_idx, col_idx).value) for row_idx in range(2, ws.max_row + 1)]
        if values and all(value == 0 for value in values):
            ws.delete_cols(col_idx)


def display_value(value, number_format):
    if value is None:
        return ""
    if not isinstance(value, (int, float)):
        return str(value)
    if "#" in number_format and "," in number_format:
        if "." in number_format:
            return f"{value:,.2f}"
        return f"{value:,.0f}"
    if "." in number_format:
        return f"{value:.2f}"
    return f"{value:.0f}"


def load_font(size=18, bold=False):
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def render_summary_table_image(ws):
    scale = 2
    cell_w = 190 * scale
    first_w = 300 * scale
    row_h = 48 * scale
    widths = [first_w] + [cell_w] * (ws.max_column - 1)
    width = sum(widths)
    height = row_h * ws.max_row
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    font = load_font(20 * scale)
    bold_font = load_font(20 * scale, True)
    header_fill = "#D8C1E8"
    metric_fill = "#E7D9F2"
    group_fills = [
        (2, 5, "#F7E6D5"),
        (6, 11, "#E4F0EA"),
        (12, 15, "#E5EDF8"),
        (16, 18, "#F3E7F7"),
        (19, 21, "#FEF0D7"),
        (22, 25, "#E9E2D5"),
    ]
    x_positions = [0]
    for width_item in widths[:-1]:
        x_positions.append(x_positions[-1] + width_item)
    for row_idx in range(1, ws.max_row + 1):
        y = (row_idx - 1) * row_h
        if row_idx == 1:
            row_fill = header_fill
        else:
            row_fill = "#FFFFFF"
            for start, end, color in group_fills:
                if start <= row_idx <= end:
                    row_fill = color
                    break
        for col_idx in range(1, ws.max_column + 1):
            x = x_positions[col_idx - 1]
            w = widths[col_idx - 1]
            fill = header_fill if row_idx == 1 else (metric_fill if col_idx == 1 else row_fill)
            draw.rectangle([x, y, x + w, y + row_h], fill=fill, outline="#D8D0C7")
            cell = ws.cell(row_idx, col_idx)
            text = display_value(cell.value, cell.number_format)
            text_font = bold_font if row_idx == 1 or col_idx == 1 else font
            bbox = draw.textbbox((0, 0), text, font=text_font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
            pad = 14 * scale
            tx = x + pad if col_idx == 1 else x + w - text_w - pad
            if row_idx == 1:
                tx = x + max(pad, (w - text_w) / 2)
            ty = y + (row_h - text_h) / 2 - 1
            draw.text((tx, ty), text, fill="#141413", font=text_font)
    return image


def stitch_screenshots(image_paths):
    images = []
    for image_path in image_paths:
        try:
            image = Image.open(image_path).convert("RGB")
        except Exception:
            continue
        images.append(image)
    if not images:
        return None
    target_width = max(image.width for image in images)
    normalized = []
    for image in images:
        if image.width != target_width:
            ratio = target_width / image.width
            image = image.resize((target_width, int(image.height * ratio)), Image.Resampling.LANCZOS)
        normalized.append(image)
    total_height = sum(image.height for image in normalized)
    stitched = Image.new("RGB", (target_width, total_height), "white")
    y = 0
    for image in normalized:
        stitched.paste(image, (0, y))
        y += image.height
    return stitched


def combined_summary_image(ws, image_paths):
    screenshots = stitch_screenshots(image_paths)
    if screenshots is None:
        return None
    table = render_summary_table_image(ws)
    target_height = table.height
    if screenshots.height != target_height:
        ratio = target_height / screenshots.height
        screenshots = screenshots.resize((max(1, int(screenshots.width * ratio)), target_height), Image.Resampling.LANCZOS)
    combined = Image.new("RGB", (table.width + screenshots.width, target_height), "white")
    combined.paste(table, (0, 0))
    combined.paste(screenshots, (table.width, 0))
    return combined


def insert_summary_images(ws, image_paths, combined_image_path=None):
    if not image_paths:
        return None
    combined = combined_summary_image(ws, image_paths)
    if combined is None:
        return None
    output_path = Path(combined_image_path) if combined_image_path else Path(image_paths[0]).with_name(f"{Path(image_paths[0]).stem}_summary_combined.png")
    combined.save(output_path)
    image = ExcelImage(str(output_path))
    max_width = 1400
    if image.width and image.width > max_width:
        ratio = max_width / image.width
        image.width = max_width
        image.height = int(image.height * ratio)
    ws.add_image(image, f"A{ws.max_row + 3}")
    return output_path


def populate_source(ws, rows):
    ws.append(SOURCE_HEADERS)
    for raw in rows:
        row = []
        for header in SOURCE_HEADERS:
            value = raw.get(header, "")
            if header in NUMERIC_HEADERS:
                row.append(to_number(value))
            elif header in {"创建时间", "加热开始时间", "加热结束时间"}:
                row.append(normalize_date_text(value))
            elif str(value).strip().upper() == "N/A":
                row.append("")
            else:
                row.append(value)
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[3].number_format = "@"
        for cell in row[6:]:
            cell.number_format = "0.00" if isinstance(cell.value, float) and not cell.value.is_integer() else "0"
    style_sheet(ws)
    set_widths(ws, [20, 14, 24, 20, 14, 14, 18, 12, 16, 14, 14, 14, 14, 20, 14, 12, 14, 14, 14])


def sum_formula(target, col_letter, last_row):
    data_range = f"'直播投放数据源'!${col_letter}$2:${col_letter}${last_row}"
    name_range = f"'直播投放数据源'!$C$2:$C${last_row}"
    if target == "":
        return f"=SUM({data_range})"
    return f'=SUMIF({name_range},"*{target}*",{data_range})'


def safe_div(numerator, denominator):
    return f"=IFERROR({numerator}/{denominator},0)"


def numeric_format(use_thousands=True, decimal_mode="fixed2"):
    prefix = "#,##" if use_thousands else ""
    if decimal_mode == "full":
        return f"{prefix}0.##########"
    return f"{prefix}0.00"


def integer_format(use_thousands=True):
    return "#,##0" if use_thousands else "0"


def percent_format(decimal_mode="fixed2"):
    if decimal_mode == "full":
        return "0.##########%"
    return "0.00%"


def ratio_display_format(metric, decimal_mode="fixed2", rate_metrics_as_percent=True):
    percent_metrics = {"直播间进入率", "新增/场观", "评论/场观"}
    if rate_metrics_as_percent and metric in percent_metrics:
        return percent_format(decimal_mode)
    return numeric_format(False, decimal_mode)


def populate_summary(
    ws,
    source_last_row,
    use_thousands=True,
    decimal_mode="fixed2",
    transpose=True,
    aggregates=None,
    remove_zero_columns=False,
    rate_metrics_as_percent=True,
):
    if not transpose:
        populate_summary_horizontal(ws, source_last_row, use_thousands, decimal_mode, aggregates, rate_metrics_as_percent)
        return

    ws.append(["指标"] + ["全部"] + TARGET_ROWS[1:])
    metric_rows = {}
    for row_idx, metric in enumerate(SUMMARY_METRICS, 2):
        ws.cell(row_idx, 1, metric)
        metric_rows[metric] = row_idx

    source_cols = {
        "总消耗": "H",
        "直播间曝光人数": "I",
        "直播间观看人数": "J",
        "直播间新增粉丝数": "K",
        "直播间评论次数": "L",
        "直播间点赞次数": "M",
        "直播间超过1分钟观看人数": "N",
        "总成交金额": "O",
        "总成交订单数": "Q",
        "总下单金额": "R",
    }

    for col_idx, target in enumerate(TARGET_ROWS, 2):
        if aggregates:
            values = summary_values(aggregates, target)
            for metric, value in values.items():
                ws.cell(metric_rows[metric], col_idx, value)
        else:
            col = get_column_letter(col_idx)
            ws.cell(metric_rows["总消耗（豆）"], col_idx, sum_formula(target, source_cols["总消耗"], source_last_row))
            ws.cell(metric_rows["总消耗金额"], col_idx, f"={col}{metric_rows['总消耗（豆）']}/10")
            ws.cell(metric_rows["服务费"], col_idx, f"={col}{metric_rows['总消耗金额']}*10%")
            ws.cell(metric_rows["合计金额"], col_idx, f"={col}{metric_rows['总消耗金额']}+{col}{metric_rows['服务费']}")
            ws.cell(metric_rows["总下单金额"], col_idx, sum_formula(target, source_cols["总下单金额"], source_last_row))
            ws.cell(metric_rows["下单roi"], col_idx, safe_div(f"{col}{metric_rows['总下单金额']}", f"{col}{metric_rows['总消耗金额']}"))
            ws.cell(metric_rows["总成交金额"], col_idx, sum_formula(target, source_cols["总成交金额"], source_last_row))
            ws.cell(metric_rows["总成交roi"], col_idx, safe_div(f"{col}{metric_rows['总成交金额']}", f"{col}{metric_rows['总消耗金额']}"))
            ws.cell(metric_rows["总成交订单数"], col_idx, sum_formula(target, source_cols["总成交订单数"], source_last_row))
            ws.cell(metric_rows["单个成交成本"], col_idx, safe_div(f"{col}{metric_rows['总消耗金额']}", f"{col}{metric_rows['总成交订单数']}"))
            ws.cell(metric_rows["曝光人数"], col_idx, sum_formula(target, source_cols["直播间曝光人数"], source_last_row))
            ws.cell(metric_rows["观看人数"], col_idx, sum_formula(target, source_cols["直播间观看人数"], source_last_row))
            ws.cell(metric_rows["场观均价"], col_idx, safe_div(f"{col}{metric_rows['总消耗金额']}", f"{col}{metric_rows['观看人数']}"))
            ws.cell(metric_rows["直播间进入率"], col_idx, safe_div(f"{col}{metric_rows['观看人数']}", f"{col}{metric_rows['曝光人数']}"))
            ws.cell(metric_rows["新增粉丝"], col_idx, sum_formula(target, source_cols["直播间新增粉丝数"], source_last_row))
            ws.cell(metric_rows["新增均价"], col_idx, safe_div(f"{col}{metric_rows['总消耗金额']}", f"{col}{metric_rows['新增粉丝']}"))
            ws.cell(metric_rows["新增/场观"], col_idx, safe_div(f"{col}{metric_rows['新增粉丝']}", f"{col}{metric_rows['观看人数']}"))
            ws.cell(metric_rows["评论次数"], col_idx, sum_formula(target, source_cols["直播间评论次数"], source_last_row))
            ws.cell(metric_rows["评论均价"], col_idx, safe_div(f"{col}{metric_rows['总消耗金额']}", f"{col}{metric_rows['评论次数']}"))
            ws.cell(metric_rows["评论/场观"], col_idx, safe_div(f"{col}{metric_rows['评论次数']}", f"{col}{metric_rows['观看人数']}"))
            ws.cell(metric_rows["点赞次数"], col_idx, sum_formula(target, source_cols["直播间点赞次数"], source_last_row))
            ws.cell(metric_rows["互动总数"], col_idx, f"={col}{metric_rows['评论次数']}+{col}{metric_rows['点赞次数']}")
            ws.cell(metric_rows["超过1分钟观看人数"], col_idx, sum_formula(target, source_cols["直播间超过1分钟观看人数"], source_last_row))
            ws.cell(metric_rows["一分钟观看均价"], col_idx, safe_div(f"{col}{metric_rows['总消耗金额']}", f"{col}{metric_rows['超过1分钟观看人数']}"))

    integer_metrics = {
        "总消耗（豆）",
        "总成交订单数",
        "曝光人数",
        "观看人数",
        "新增粉丝",
        "评论次数",
        "点赞次数",
        "互动总数",
        "超过1分钟观看人数",
    }
    percent_metrics = {"下单roi", "总成交roi", "直播间进入率", "新增/场观", "评论/场观"}
    for row_idx in range(2, ws.max_row + 1):
        metric = ws.cell(row_idx, 1).value
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if metric in percent_metrics:
                cell.number_format = ratio_display_format(metric, decimal_mode, rate_metrics_as_percent)
            elif metric in integer_metrics:
                cell.number_format = integer_format(use_thousands)
            else:
                cell.number_format = numeric_format(use_thousands, decimal_mode)

    if remove_zero_columns:
        delete_zero_summary_columns(ws)
    style_sheet(ws, "7030A0")
    apply_summary_group_styles(ws, True)
    set_widths(ws, [20] + [14] * (ws.max_column - 1))


def populate_summary_horizontal(
    ws,
    source_last_row,
    use_thousands=True,
    decimal_mode="fixed2",
    aggregates=None,
    rate_metrics_as_percent=True,
):
    ws.append(SUMMARY_HEADERS)
    source_cols = {
        "总消耗": "H",
        "直播间曝光人数": "I",
        "直播间观看人数": "J",
        "直播间新增粉丝数": "K",
        "直播间评论次数": "L",
        "直播间点赞次数": "M",
        "直播间超过1分钟观看人数": "N",
        "总成交金额": "O",
        "总成交订单数": "Q",
        "总下单金额": "R",
    }
    for row_idx, target in enumerate(TARGET_ROWS, 2):
        ws.cell(row_idx, 1, target)
        if aggregates:
            values = summary_values(aggregates, target)
            for col_idx, metric in enumerate(SUMMARY_METRICS, 2):
                ws.cell(row_idx, col_idx, values[metric])
        else:
            ws.cell(row_idx, 2, sum_formula(target, source_cols["总消耗"], source_last_row))
            ws.cell(row_idx, 3, f"=B{row_idx}/10")
            ws.cell(row_idx, 4, f"=C{row_idx}*10%")
            ws.cell(row_idx, 5, f"=C{row_idx}+D{row_idx}")
            ws.cell(row_idx, 6, sum_formula(target, source_cols["总下单金额"], source_last_row))
            ws.cell(row_idx, 7, safe_div(f"F{row_idx}", f"C{row_idx}"))
            ws.cell(row_idx, 8, sum_formula(target, source_cols["总成交金额"], source_last_row))
            ws.cell(row_idx, 9, safe_div(f"H{row_idx}", f"C{row_idx}"))
            ws.cell(row_idx, 10, sum_formula(target, source_cols["总成交订单数"], source_last_row))
            ws.cell(row_idx, 11, safe_div(f"C{row_idx}", f"J{row_idx}"))
            ws.cell(row_idx, 12, sum_formula(target, source_cols["直播间曝光人数"], source_last_row))
            ws.cell(row_idx, 13, sum_formula(target, source_cols["直播间观看人数"], source_last_row))
            ws.cell(row_idx, 14, safe_div(f"C{row_idx}", f"M{row_idx}"))
            ws.cell(row_idx, 15, safe_div(f"M{row_idx}", f"L{row_idx}"))
            ws.cell(row_idx, 16, sum_formula(target, source_cols["直播间新增粉丝数"], source_last_row))
            ws.cell(row_idx, 17, safe_div(f"C{row_idx}", f"P{row_idx}"))
            ws.cell(row_idx, 18, safe_div(f"P{row_idx}", f"M{row_idx}"))
            ws.cell(row_idx, 19, sum_formula(target, source_cols["直播间评论次数"], source_last_row))
            ws.cell(row_idx, 20, safe_div(f"C{row_idx}", f"S{row_idx}"))
            ws.cell(row_idx, 21, safe_div(f"S{row_idx}", f"M{row_idx}"))
            ws.cell(row_idx, 22, sum_formula(target, source_cols["直播间点赞次数"], source_last_row))
            ws.cell(row_idx, 23, f"=S{row_idx}+V{row_idx}")
            ws.cell(row_idx, 24, sum_formula(target, source_cols["直播间超过1分钟观看人数"], source_last_row))
            ws.cell(row_idx, 25, safe_div(f"C{row_idx}", f"X{row_idx}"))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, 1):
            if idx in {7, 9, 15, 18, 21}:
                metric = SUMMARY_HEADERS[idx - 1]
                cell.number_format = ratio_display_format(metric, decimal_mode, rate_metrics_as_percent)
            elif idx in {3, 4, 5, 6, 8, 11, 14, 17, 20, 25}:
                cell.number_format = numeric_format(use_thousands, decimal_mode)
            else:
                cell.number_format = integer_format(use_thousands)
    style_sheet(ws, "7030A0")
    set_widths(ws, [12, 14, 14, 12, 14, 14, 12, 14, 12, 14, 14, 12, 12, 12, 14, 12, 12, 12, 12, 12, 12, 12, 12, 18, 16])


def summary_ref(metric, target_col="B"):
    row_idx = SUMMARY_METRICS.index(metric) + 2
    return f"='数据汇总'!{target_col}{row_idx}"


def populate_settlement(ws, transpose=True, aggregates=None, first_date=""):
    ws.append(SETTLEMENT_HEADERS)
    if aggregates:
        values = summary_values(aggregates, "")
        formulas = [
            first_date,
            values["总成交金额"],
            values["总成交roi"],
            values["总成交订单数"],
            values["单个成交成本"],
            values["曝光人数"],
            values["观看人数"],
            values["场观均价"],
            values["直播间进入率"],
            values["新增粉丝"],
            values["新增均价"],
            values["评论次数"],
            values["总消耗（豆）"],
            values["总消耗金额"],
            values["点赞次数"],
        ]
    elif transpose:
        formulas = [
            "=INDEX('直播投放数据源'!B:B,2)",
            summary_ref("总成交金额"),
            summary_ref("总成交roi"),
            summary_ref("总成交订单数"),
            summary_ref("单个成交成本"),
            summary_ref("曝光人数"),
            summary_ref("观看人数"),
            summary_ref("场观均价"),
            summary_ref("直播间进入率"),
            summary_ref("新增粉丝"),
            summary_ref("新增均价"),
            summary_ref("评论次数"),
            summary_ref("总消耗（豆）"),
            summary_ref("总消耗金额"),
            summary_ref("点赞次数"),
        ]
    else:
        formulas = [
            "=INDEX('直播投放数据源'!B:B,2)",
            "='数据汇总'!H2",
            "='数据汇总'!I2",
            "='数据汇总'!J2",
            "='数据汇总'!K2",
            "='数据汇总'!L2",
            "='数据汇总'!M2",
            "='数据汇总'!N2",
            "='数据汇总'!O2",
            "='数据汇总'!P2",
            "='数据汇总'!Q2",
            "='数据汇总'!S2",
            "='数据汇总'!B2",
            "='数据汇总'!C2",
            "='数据汇总'!V2",
        ]
    ws.append(formulas)
    money_cols = {2, 5, 8, 11, 13, 14}
    pct_cols = {3, 9}
    comma_cols = {6, 7, 10, 12, 15}
    for idx, cell in enumerate(ws[2], 1):
        if idx in money_cols:
            cell.number_format = '¥#,##0.00'
        elif idx in pct_cols:
            cell.number_format = "0.00%"
        elif idx in comma_cols:
            cell.number_format = "#,##0"
        else:
            cell.number_format = "0"
    ws["A2"].number_format = "yyyy年m月d日"
    style_sheet(ws, "548235")
    set_widths(ws, [16, 12, 10, 12, 14, 12, 10, 12, 12, 10, 12, 10, 16, 18, 10])


def build_workbook(
    csv_path,
    output_path,
    use_thousands=True,
    decimal_mode="fixed2",
    transpose_summary=True,
    remove_zero_columns=False,
    image_paths=None,
    combined_image_path=None,
    rate_metrics_as_percent=True,
):
    csv_paths = [csv_path] if isinstance(csv_path, (str, Path)) else list(csv_path)
    header_sets, rows = read_csv_files(csv_paths)
    warnings = validate_inputs(header_sets, rows)
    aggregates = aggregate_rows(rows)
    first_date = (
        normalize_date_text(rows[0].get("加热开始时间", ""))
        or normalize_date_text(rows[0].get("创建时间", ""))
        if rows
        else ""
    )
    wb = Workbook()
    source = wb.active
    source.title = "直播投放数据源"
    populate_source(source, rows)
    summary = wb.create_sheet("数据汇总")
    populate_summary(
        summary,
        max(source.max_row, 2),
        use_thousands,
        decimal_mode,
        transpose_summary,
        aggregates,
        remove_zero_columns,
        rate_metrics_as_percent,
    )
    combined_image = insert_summary_images(summary, image_paths or [], combined_image_path)
    settlement = wb.create_sheet("结算整理表")
    populate_settlement(settlement, transpose_summary, aggregates, first_date)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"warnings": warnings, "default_stem": default_output_stem(rows), "combined_image": str(combined_image) if combined_image else None}


def main():
    parser = argparse.ArgumentParser(description="将直播投放 CSV 转换为三表结构 XLSX。")
    parser.add_argument("csv", nargs="+", help="输入 CSV 文件路径，可传多个")
    parser.add_argument("-o", "--output", help="输出 XLSX 文件路径")
    parser.add_argument("--image", action="append", default=[], help="插入到数据汇总旁边的截图路径，可重复")
    parser.add_argument("--no-thousands", action="store_true", help="数字不使用千分位")
    parser.add_argument("--decimal-mode", choices=["fixed2", "full"], default="fixed2", help="小数格式")
    parser.add_argument("--no-transpose-summary", action="store_true", help="数据情况横向不转置")
    parser.add_argument("--remove-zero-columns", action="store_true", help="删除数据全为 0 的分类列")
    parser.add_argument("--rate-metrics-as-decimal", action="store_true", help="进入率/新增占比/评论占比按小数显示")
    args = parser.parse_args()
    csv_paths = [Path(path) for path in args.csv]
    header_sets, rows = read_csv_files(csv_paths)
    default_stem = default_output_stem(rows)
    output_path = Path(args.output) if args.output else csv_paths[0].with_name(f"{default_stem}.xlsx")
    result = build_workbook(
        csv_paths,
        output_path,
        use_thousands=not args.no_thousands,
        decimal_mode=args.decimal_mode,
        transpose_summary=not args.no_transpose_summary,
        remove_zero_columns=args.remove_zero_columns,
        image_paths=[Path(path) for path in args.image],
        rate_metrics_as_percent=not args.rate_metrics_as_decimal,
    )
    print(output_path)
    for warning in result["warnings"]:
        print(f"警告：{warning}")


if __name__ == "__main__":
    main()
