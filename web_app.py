#!/usr/bin/env python3
import cgi
import json
import mimetypes
import os
import socket
import sys
import tempfile
import threading
import uuid
import webbrowser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote, unquote, urlparse

from openpyxl import load_workbook

from process_live_ads import (
    build_workbook,
    default_output_stem,
    read_csv_files,
    to_number,
    validate_inputs,
)


def app_base_dir():
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    return Path(__file__).resolve().parent


APP_DIR = app_base_dir()
OUTPUT_DIR = Path(tempfile.gettempdir()) / "live_ads_panel_outputs"
STATIC_DIR = APP_DIR / "static"


def summarize_rows(rows):
    totals = {
        "rows": len(rows),
        "spend_beans": sum(to_number(row.get("总消耗")) for row in rows),
        "deal_amount": sum(to_number(row.get("总成交金额")) for row in rows),
        "order_amount": sum(to_number(row.get("总下单金额")) for row in rows),
        "deal_orders": sum(to_number(row.get("总成交订单数")) for row in rows),
    }
    totals["spend_yuan"] = totals["spend_beans"] / 10
    totals["deal_roi"] = totals["deal_amount"] / totals["spend_yuan"] if totals["spend_yuan"] else 0
    return totals


def form_bool(form, name, default=True):
    if name not in form:
        return default
    value = str(form.getfirst(name, "")).lower()
    return value in {"1", "true", "yes", "on"}


def safe_download_name(name):
    stem = Path(name).stem or "投放数据处理结果"
    stem = "".join(ch for ch in stem if ch not in '\\/:*?"<>|').strip()
    return f"{stem}_处理结果.xlsx"


def safe_xlsx_name(name):
    stem = Path(name).stem or "投放数据处理结果"
    stem = "".join(ch for ch in stem if ch not in '\\/:*?"<>|').strip()
    return f"{stem}.xlsx"


def form_files(form, name):
    if name not in form:
        return []
    value = form[name]
    files = value if isinstance(value, list) else [value]
    return [item for item in files if getattr(item, "filename", "")]


def sheet_preview(workbook_path, sheet_name):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append(["" if value is None else value for value in row])
    return rows


class AppHandler(BaseHTTPRequestHandler):
    server_version = "LiveAdsPanel/1.0"

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
            return
        if parsed.path.startswith("/static/"):
            rel_path = unquote(parsed.path.removeprefix("/static/"))
            self.serve_static(rel_path)
            return
        if parsed.path.startswith("/download/"):
            self.serve_download(parsed.path.removeprefix("/download/"))
            return
        if parsed.path.startswith("/image/"):
            self.serve_image(parsed.path.removeprefix("/image/"))
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not Found")

    def do_POST(self):
        if urlparse(self.path).path != "/api/process":
            self.send_error(HTTPStatus.NOT_FOUND, "接口不存在")
            return

        content_type = self.headers.get("content-type", "")
        if "multipart/form-data" not in content_type:
            self.send_json({"ok": False, "error": "请上传 CSV 文件。"}, HTTPStatus.BAD_REQUEST)
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": self.headers.get("content-length", "0"),
            },
        )
        uploads = form_files(form, "files") or form_files(form, "file")
        if not uploads:
            self.send_json({"ok": False, "error": "没有收到文件。"}, HTTPStatus.BAD_REQUEST)
            return

        for upload in uploads:
            if not Path(upload.filename).name.lower().endswith(".csv"):
                self.send_json({"ok": False, "error": "目前只支持 CSV 文件。"}, HTTPStatus.BAD_REQUEST)
                return

        job_id = uuid.uuid4().hex
        job_dir = OUTPUT_DIR / job_id
        job_dir.mkdir(parents=True, exist_ok=True)
        temp_csvs = []
        temp_images = []
        custom_output_name = str(form.getfirst("outputName", "")).strip()
        use_thousands = form_bool(form, "useThousands", True)
        transpose_summary = form_bool(form, "transposeSummary", True)
        remove_zero_columns = form_bool(form, "removeZeroColumns", False)
        rate_metrics_as_percent = form_bool(form, "rateMetricsAsPercent", True)
        decimal_mode = str(form.getfirst("decimalMode", "fixed2"))
        if decimal_mode not in {"fixed2", "full"}:
            decimal_mode = "fixed2"

        try:
            for idx, upload in enumerate(uploads, 1):
                temp_csv = Path(tempfile.gettempdir()) / f"live_ads_{job_id}_{idx}.csv"
                with temp_csv.open("wb") as f:
                    f.write(upload.file.read())
                temp_csvs.append(temp_csv)
            screenshots = form_files(form, "screenshots")
            for idx, screenshot in enumerate(screenshots, 1):
                suffix = Path(screenshot.filename).suffix.lower() or ".png"
                temp_image = Path(tempfile.gettempdir()) / f"live_ads_{job_id}_screenshot_{idx}{suffix}"
                with temp_image.open("wb") as f:
                    f.write(screenshot.file.read())
                temp_images.append(temp_image)

            header_sets, rows = read_csv_files(temp_csvs)
            if not rows:
                raise ValueError("CSV 没有数据行。")
            warnings = validate_inputs(header_sets, rows)
            output_name = safe_xlsx_name(custom_output_name or default_output_stem(rows))
            output_path = job_dir / output_name
            combined_image_name = f"{Path(output_name).stem}_拼接图.png"
            combined_image_path = job_dir / combined_image_name
            result = build_workbook(
                temp_csvs,
                output_path,
                use_thousands=use_thousands,
                decimal_mode=decimal_mode,
                transpose_summary=transpose_summary,
                remove_zero_columns=remove_zero_columns,
                image_paths=temp_images,
                combined_image_path=combined_image_path,
                rate_metrics_as_percent=rate_metrics_as_percent,
            )
            warnings = warnings or result.get("warnings", [])
            summary = summarize_rows(rows)
            previews = {
                "sheet2": sheet_preview(output_path, "数据汇总"),
                "sheet3": sheet_preview(output_path, "结算整理表"),
            }
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        finally:
            try:
                for temp_path in temp_csvs + temp_images:
                    temp_path.unlink()
            except FileNotFoundError:
                pass

        self.send_json(
            {
                "ok": True,
                "fileName": output_name,
                "downloadUrl": f"/download/{job_id}/{quote(output_name)}",
                "summary": summary,
                "previews": previews,
                "warnings": warnings,
                "combinedImageUrl": f"/image/{job_id}/{quote(combined_image_name)}" if result.get("combined_image") else "",
                "combinedImageName": combined_image_name if result.get("combined_image") else "",
            }
        )

    def serve_file(self, path, content_type=None):
        if not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "File Not Found")
            return
        data = path.read_bytes()
        content_type = content_type or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_static(self, rel_path):
        full_path = (STATIC_DIR / rel_path).resolve()
        if STATIC_DIR.resolve() not in full_path.parents and full_path != STATIC_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        self.serve_file(full_path)

    def serve_download(self, rel_path):
        full_path = (OUTPUT_DIR / unquote(rel_path)).resolve()
        if OUTPUT_DIR.resolve() not in full_path.parents:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        if not full_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "File Not Found")
            return
        data = full_path.read_bytes()
        encoded_name = quote(full_path.name)
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(data)))
        self.send_header(
            "Content-Disposition",
            f"attachment; filename=\"output.xlsx\"; filename*=UTF-8''{encoded_name}",
        )
        self.end_headers()
        self.wfile.write(data)

    def serve_image(self, rel_path):
        full_path = (OUTPUT_DIR / unquote(rel_path)).resolve()
        if OUTPUT_DIR.resolve() not in full_path.parents:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        if not full_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "File Not Found")
            return
        data = full_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(full_path.name)[0] or "image/png")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f"attachment; filename=\"summary.png\"; filename*=UTF-8''{quote(full_path.name)}")
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        print(f"{self.address_string()} - {fmt % args}")


def main():
    port = find_available_port(int(os.environ.get("PORT", "8765")))
    server = ThreadingHTTPServer(("127.0.0.1", port), AppHandler)
    url = f"http://127.0.0.1:{port}"
    if os.environ.get("NO_BROWSER") != "1":
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    print(f"投放数据处理面板已启动：{url}")
    server.serve_forever()


def find_available_port(start_port):
    for port in range(start_port, start_port + 50):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                sock.bind(("127.0.0.1", port))
            except OSError:
                continue
            return port
    raise RuntimeError("没有找到可用端口。")


if __name__ == "__main__":
    main()
